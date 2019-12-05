from keras.models import Sequential
from keras.layers import Dense
from sklearn.metrics import mean_squared_error
from sklearn.metrics import r2_score
from openpyxl import load_workbook
import numpy as np
load_xl = load_workbook("\Study\Keras\lotto.xlsx", data_only=True)

load_sheet = load_xl['lotto']

lotto_round = np.array([int(load_sheet.cell(x,2).value) for x in range(890, 5, -1)])
lotto_num = np.array([[int(load_sheet.cell(i,j).value) for j in range(14,20)] for i in range(890, 5, -1)])

lotto_data1 = np.array([int(load_sheet.cell(x,2).value) for x in range(890, 5, -1)])
lotto_data2 = np.array([[int(load_sheet.cell(i,j).value) for j in range(14,20)] for i in range(890, 5, -1)])

model = Sequential()
model.add(Dense(200, input_dim=1, activation='relu'))


for i in range(161, 0, -40):
    model.add(Dense(i))
model.add(Dense(6))

model.summary()

model.compile(loss='mse', optimizer='adam', metrics=['accuracy']) # accuracy
model.fit(lotto_round, lotto_num, epochs=100, batch_size=1)

loss, acc = model.evaluate(lotto_data1, lotto_data2, batch_size=1)

print("acc :", acc)
print("loss : ", loss)

y_predict = model.predict([887])
print(y_predict)